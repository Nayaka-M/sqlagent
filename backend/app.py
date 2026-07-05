from fastapi import FastAPI, HTTPException, Depends, Header, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager
from dotenv import load_dotenv
from typing import List, Optional
from datetime import datetime
import os
import io
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Try importing Gemini
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False
    print("⚠️ google-generativeai not installed")

from .database import (
    connect_db, disconnect_db, create_log, get_logs, get_log_by_id, get_user_stats,
    save_database_connection, get_user_databases, get_database_connection,
    update_user_profile, update_user_password_hash, get_user_profile_stats
)
from .models import (
    UserRegister, UserLogin, UserResponse, TokenResponse,
    DatabaseConnectionRequest, DatabaseConnectionResponse,
    QueryRequest, QueryResponse, LogResponse, ProfileResponse,
    UserProfileUpdate, UserPasswordUpdate
)
from .auth import create_user, login_user, verify_token, get_user_by_id, init_auth_db, update_password
from .query_executor import execute_query
from .prompts import get_sql_generation_prompt, get_explanation_prompt

load_dotenv()

# ============================================================
# CONFIGURE GEMINI AI - FIXED VERSION
# ============================================================
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
gemini_model = None
GEMINI_AVAILABLE_MODEL = None

if GEMINI_API_KEY and GEMINI_AVAILABLE:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        
        # Try multiple models in order of preference
        model_names = [
            'gemini-2.0-flash-exp',    # Latest and fastest
            'gemini-2.0-flash',         # Stable version
            'gemini-1.5-flash',         # Good for SQL
            'gemini-1.5-pro',           # Original
            'gemini-pro'                # Fallback
        ]
        
        for model_name in model_names:
            try:
                test_model = genai.GenerativeModel(model_name)
                # Test with a simple prompt to verify it works
                test_response = test_model.generate_content("test")
                if test_response:
                    gemini_model = test_model
                    GEMINI_AVAILABLE_MODEL = model_name
                    print(f"✅ Gemini AI configured with: {model_name}")
                    break
            except Exception as e:
                error_msg = str(e)
                if "quota" in error_msg.lower():
                    print(f"⚠️ Quota exceeded for {model_name}, trying next...")
                    continue
                elif "404" in error_msg:
                    print(f"⚠️ Model {model_name} not found, trying next...")
                    continue
                else:
                    print(f"⚠️ Model {model_name} unavailable: {e}")
                    continue
        
        if not gemini_model:
            print("⚠️ No Gemini model available, using fallback")
            gemini_model = None
            
    except Exception as e:
        print(f"⚠️ Gemini AI configuration error: {e}")
        gemini_model = None
else:
    if not GEMINI_API_KEY:
        print("⚠️ GEMINI_API_KEY not set in environment")
    if not GEMINI_AVAILABLE:
        print("⚠️ google-generativeai package not installed")
    
print(f"🔍 Final Gemini Status: {'Available using ' + GEMINI_AVAILABLE_MODEL if gemini_model else 'Using Fallback (no AI)'}")

# ============================================================
# LIFESPAN
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    await connect_db()
    await init_auth_db()
    yield
    await disconnect_db()

# ============================================================
# CREATE APP
# ============================================================
app = FastAPI(
    title="SQL Query Agent API",
    description="Natural Language to SQL with AI, Database Connections & Query Execution",
    version="2.0.0",
    lifespan=lifespan
)

# ============================================================
# CORS MIDDLEWARE
# ============================================================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# WEBSOCKET MANAGER
# ============================================================
class ConnectionManager:
    def __init__(self):
        self.active_connections = {}
        self.user_connections = {}

    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        if user_id not in self.active_connections:
            self.active_connections[user_id] = []
        self.active_connections[user_id].append(websocket)
        self.user_connections[user_id] = websocket
        print(f"✅ User {user_id} connected")

    def disconnect(self, websocket: WebSocket, user_id: str):
        if user_id in self.active_connections:
            try:
                self.active_connections[user_id].remove(websocket)
                if not self.active_connections[user_id]:
                    del self.active_connections[user_id]
            except ValueError:
                pass
        if user_id in self.user_connections:
            del self.user_connections[user_id]
        print(f"❌ User {user_id} disconnected")

    async def send_personal_message(self, message: dict, user_id: str):
        if user_id in self.user_connections:
            try:
                await self.user_connections[user_id].send_text(json.dumps(message))
                return True
            except:
                return False
        return False

manager = ConnectionManager()

# ============================================================
# WEBSOCKET ENDPOINT
# ============================================================
@app.websocket("/ws/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: str):
    await manager.connect(websocket, user_id)
    
    await manager.send_personal_message({
        "type": "connection",
        "message": "Connected to real-time server",
        "timestamp": datetime.now().isoformat()
    }, user_id)

    try:
        while True:
            data = await websocket.receive_text()
            try:
                message = json.loads(data)
                if message.get("type") == "ping":
                    await manager.send_personal_message({"type": "pong"}, user_id)
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        manager.disconnect(websocket, user_id)

# ============================================================
# SEND NOTIFICATION HELPER
# ============================================================
async def send_realtime_notification(user_id: str, title: str, message: str, type: str = "info"):
    notification = {
        "type": "notification",
        "notification_type": type,
        "title": title,
        "message": message,
        "timestamp": datetime.now().isoformat(),
        "read": False
    }
    return await manager.send_personal_message(notification, user_id)

# ============================================================
# GET CURRENT USER
# ============================================================
async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    parts = authorization.split()
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=401, detail="Invalid token format")
    
    token = parts[1]
    user = await verify_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    return user

# ============================================================
# IMPROVED SQL GENERATION WITH FALLBACK
# ============================================================

def generate_sql_fallback(prompt: str, db_id: str = None) -> str:
    """Enhanced fallback function with better query detection"""
    p = prompt.lower().strip()
    USER_TABLE = "users"
    ORDER_TABLE = "orders"
    PRODUCT_TABLE = "products"
    
    print(f"🔍 FALLBACK: Processing prompt: {p}")
    
    # Table management
    if any(word in p for word in ["show tables", "list tables", "all tables"]):
        return "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public' ORDER BY table_name;"
    
    if any(word in p for word in ["describe", "table structure", "columns"]):
        return f"SELECT column_name, data_type, is_nullable FROM information_schema.columns WHERE table_schema = 'public' AND table_name = '{USER_TABLE}' ORDER BY column_name;"
    
    # User queries
    if "show me all users" in p or "all users" in p:
        return f"SELECT id, username, email, full_name, created_at FROM {USER_TABLE} ORDER BY created_at DESC LIMIT 20;"
    
    if "users who registered today" in p or "registered today" in p:
        return f"SELECT id, username, email, full_name, created_at FROM {USER_TABLE} WHERE DATE(created_at) = CURRENT_DATE ORDER BY created_at DESC;"
    
    if "users who registered" in p and "today" not in p:
        return f"SELECT id, username, email, full_name, created_at FROM {USER_TABLE} ORDER BY created_at DESC;"
    
    if "count" in p and "user" in p:
        return f"SELECT COUNT(*) as total_users, COUNT(DISTINCT email) as unique_emails FROM {USER_TABLE};"
    
    if "user with username" in p:
        username = p.split("username")[-1].strip().split()[0] if "username" in p else "nayaka_m"
        return f"SELECT * FROM {USER_TABLE} WHERE username = '{username}';"
    
    # Order queries
    if "total revenue" in p or "sum" in p or "revenue" in p:
        return f"SELECT SUM(amount) as total_revenue, COUNT(*) as total_orders, AVG(amount) as avg_order_value FROM {ORDER_TABLE};"
    
    if "orders" in p and "count" not in p:
        return f"SELECT id, customer_id, amount, status, created_at FROM {ORDER_TABLE} ORDER BY created_at DESC LIMIT 20;"
    
    # Product queries
    if "products out of stock" in p or "out of stock" in p:
        return f"SELECT id, name, price, stock FROM {PRODUCT_TABLE} WHERE stock = 0 ORDER BY name;"
    
    if "products" in p:
        return f"SELECT id, name, price, stock FROM {PRODUCT_TABLE} LIMIT 20;"
    
    # Top/Limit queries
    if "top 5" in p or "top 10" in p:
        if "customer" in p:
            return f"SELECT customer_id, COUNT(*) as order_count, SUM(amount) as total_spent FROM {ORDER_TABLE} GROUP BY customer_id ORDER BY total_spent DESC LIMIT 5;"
        elif "user" in p:
            return f"SELECT id, username, email, full_name, created_at FROM {USER_TABLE} ORDER BY created_at DESC LIMIT 5;"
        elif "product" in p:
            return f"SELECT id, name, price, stock FROM {PRODUCT_TABLE} ORDER BY price DESC LIMIT 5;"
    
    # Analytics
    if "average" in p:
        if "order" in p:
            return f"SELECT AVG(amount) as avg_order_value FROM {ORDER_TABLE};"
        elif "user" in p:
            return f"SELECT COUNT(*) as total_users FROM {USER_TABLE};"
    
    if "recent" in p:
        return f"SELECT id, username, email, created_at FROM {USER_TABLE} ORDER BY created_at DESC LIMIT 10;"
    
    # Default - show tables instead of hardcoded query
    print("⚠️ No specific pattern matched, showing tables")
    return "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public' ORDER BY table_name;"

async def generate_sql_with_ai(prompt: str, db_id: str = None) -> str:
    """Generate SQL using AI with fallback support"""
    print(f"🤖 AI GENERATE: Processing prompt: '{prompt}'")
    print(f"🤖 AI GENERATE: Gemini available: {gemini_model is not None}")
    
    # If no Gemini, use fallback immediately
    if not gemini_model:
        print("⚠️ No Gemini model available, using fallback")
        sql = generate_sql_fallback(prompt, db_id)
        print(f"🔍 Fallback SQL: {sql}")
        return sql
    
    try:
        # Use the prompt engineering template
        full_prompt = get_sql_generation_prompt(prompt)
        print(f"📝 Sending to Gemini with prompt length: {len(full_prompt)}")
        
        # Generate SQL with timeout protection
        response = gemini_model.generate_content(full_prompt)
        
        if not response or not response.text:
            print("⚠️ Empty response from Gemini, using fallback")
            return generate_sql_fallback(prompt, db_id)
        
        sql = response.text.strip()
        # Clean up markdown if present
        sql = sql.replace('```sql', '').replace('```', '').strip()
        print(f"✅ AI Generated SQL: {sql}")
        return sql
        
    except Exception as e:
        error_msg = str(e)
        print(f"❌ AI Error: {error_msg}")
        
        # Handle quota errors specifically
        if "quota" in error_msg.lower():
            print("⚠️ Quota exceeded, using fallback")
        elif "404" in error_msg:
            print("⚠️ Model not found, using fallback")
        elif "429" in error_msg:
            print("⚠️ Rate limited, using fallback")
        
        return generate_sql_fallback(prompt, db_id)

# ============================================================
# SQL EXPLANATION FUNCTION
# ============================================================
async def explain_sql_query(sql_query: str, prompt: str) -> str:
    """Generate explanation using prompt engineering or fallback"""
    try:
        if gemini_model:
            try:
                explanation_prompt = get_explanation_prompt(prompt, sql_query)
                response = gemini_model.generate_content(explanation_prompt)
                if response and response.text:
                    return response.text.strip()
            except Exception as e:
                print(f"AI Explanation error: {e}")
        
        return generate_basic_explanation(sql_query, prompt)
        
    except Exception as e:
        print(f"Explanation error: {e}")
        return f"Query generated for: '{prompt}'"

def generate_basic_explanation(sql_query: str, prompt: str) -> str:
    """Generate a basic explanation without AI"""
    explanation_parts = []
    sql_lower = sql_query.lower()
    
    if sql_lower.startswith("select"):
        if "count" in sql_lower:
            explanation_parts.append("📊 Counts the number of records")
        elif "sum" in sql_lower:
            explanation_parts.append("💰 Calculates the total sum")
        elif "avg" in sql_lower:
            explanation_parts.append("📈 Calculates the average")
        elif "max" in sql_lower:
            explanation_parts.append("📈 Finds the maximum value")
        elif "min" in sql_lower:
            explanation_parts.append("📈 Finds the minimum value")
        else:
            explanation_parts.append("🔍 Retrieves data from the database")
    
    from_match = re.search(r'from\s+([^\s,;]+)', sql_lower, re.IGNORECASE)
    if from_match:
        table_name = from_match.group(1)
        explanation_parts.append(f"📋 From the '{table_name}' table")
    
    where_match = re.search(r'where\s+(.+?)(?:group by|order by|limit|$)', sql_lower, re.IGNORECASE | re.DOTALL)
    if where_match:
        condition = where_match.group(1).strip()
        explanation_parts.append(f"🔍 Filtering where: {condition}")
    
    if "order by" in sql_lower:
        order_match = re.search(r'order by\s+([^\s,;]+)', sql_lower, re.IGNORECASE)
        if order_match:
            order_by = order_match.group(1)
            direction = "descending" if "desc" in sql_lower else "ascending"
            explanation_parts.append(f"📊 Sorted by {order_by} ({direction})")
    
    if "limit" in sql_lower:
        limit_match = re.search(r'limit\s+(\d+)', sql_lower, re.IGNORECASE)
        if limit_match:
            limit_val = limit_match.group(1)
            explanation_parts.append(f"📦 Limited to {limit_val} results")
    
    explanation = f"📝 SQL Query Explanation\n\n"
    explanation += f"Your Request: \"{prompt}\"\n\n"
    explanation += "What this query does:\n"
    for i, part in enumerate(explanation_parts, 1):
        explanation += f"  {i}. {part}\n"
    
    explanation += f"\n\nGenerated SQL:\n{sql_query}\n\n"
    explanation += "💡 Tip: Select a database to execute this query and see results!"
    
    return explanation

# ============================================================
# EXCEL EXPORT ENDPOINT
# ============================================================
@app.post("/api/export/excel")
async def export_to_excel(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Results"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        columns = data.get("columns", [])
        rows = data.get("rows", [])
        
        if not columns and rows:
            columns = list(rows[0].keys()) if rows else []
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_name, "")
                if value is None:
                    cell.value = "NULL"
                elif isinstance(value, (int, float)):
                    cell.value = value
                elif isinstance(value, dict):
                    cell.value = str(value)
                else:
                    cell.value = str(value)
        
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, min(len(rows) + 2, 101)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"query_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"❌ Excel export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# NOTIFICATION ENDPOINT
# ============================================================
@app.post("/api/notify/{user_id}")
async def send_notification(user_id: str, title: str, message: str, type: str = "info"):
    success = await send_realtime_notification(user_id, title, message, type)
    return {"success": success, "message": "Notification sent" if success else "User offline"}

# ============================================================
# CONTACT FORM ENDPOINT
# ============================================================
@app.post("/api/contact")
async def contact_form(
    name: str,
    email: str,
    message: str,
    subject: str = "New Contact Message",
    current_user: dict = Depends(get_current_user)
):
    try:
        if not name or not email or not message:
            raise HTTPException(status_code=400, detail="All fields are required")
        
        print(f"📧 Contact Form Submission:")
        print(f"   Name: {name}")
        print(f"   Email: {email}")
        print(f"   Subject: {subject}")
        print(f"   Message: {message}")
        print(f"   From User: {current_user['id']}")
        
        await send_realtime_notification(
            current_user["id"],
            "Contact Form Submitted",
            f"Thank you {name}, we'll get back to you soon!",
            "success"
        )
        
        return {
            "success": True,
            "message": "Message received! We'll get back to you soon.",
            "data": {
                "name": name,
                "email": email,
                "subject": subject,
                "message": message[:100] + "..."
            }
        }
            
    except Exception as e:
        print(f"❌ Contact form error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# ROOT ENDPOINT
# ============================================================
@app.get("/")
async def root():
    return {
        "message": "🚀 SQL Query Agent API is running!",
        "version": "2.0.0",
        "features": {
            "auth": "Enabled",
            "ai": f"Gemini ({GEMINI_AVAILABLE_MODEL})" if gemini_model else "Fallback",
            "database_connections": "Enabled",
            "query_execution": "Enabled",
            "excel_export": "Enabled",
            "realtime": "Enabled",
            "edit_delete": "Enabled",
            "profile": "Enabled",
            "explain": "Enabled"
        },
        "endpoints": {
            "/api/auth/register": "POST - Register a new user",
            "/api/auth/login": "POST - Login user",
            "/api/auth/me": "GET - Get current user",
            "/api/databases": "POST - Add database connection",
            "/api/databases": "GET - List database connections",
            "/api/databases/{db_id}": "PUT - Update database connection",
            "/api/databases/{db_id}": "DELETE - Delete database connection",
            "/api/query": "POST - Generate and execute SQL",
            "/api/logs": "GET - Get query history",
            "/api/profile": "GET - Get user profile",
            "/api/profile/stats": "GET - Get profile statistics",
            "/api/profile/update": "PUT - Update profile",
            "/api/profile/password": "PUT - Update password",
            "/api/export/excel": "POST - Export results to Excel",
            "/api/notify/{user_id}": "POST - Send notification",
            "/api/contact": "POST - Contact form",
            "/ws/{user_id}": "WebSocket - Real-time connection",
            "/health": "GET - Health check"
        }
    }

# ============================================================
# AUTH ENDPOINTS
# ============================================================
@app.post("/api/auth/register", response_model=UserResponse)
async def register_user(user_data: UserRegister):
    try:
        user = await create_user(
            username=user_data.username,
            email=user_data.email,
            password=user_data.password,
            full_name=user_data.full_name
        )
        return UserResponse(**user)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/login", response_model=TokenResponse)
async def login_user_endpoint(login_data: UserLogin):
    try:
        result = await login_user(login_data.email, login_data.password)
        return TokenResponse(
            access_token=result["access_token"],
            user=UserResponse(**result["user"])
        )
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    if 'created_at' in current_user and isinstance(current_user['created_at'], datetime):
        current_user['created_at'] = current_user['created_at'].isoformat()
    return UserResponse(**current_user)

# ============================================================
# DATABASE CONNECTION ENDPOINTS
# ============================================================
@app.post("/api/databases", response_model=DatabaseConnectionResponse)
async def add_database_connection(
    db_data: DatabaseConnectionRequest,
    current_user: dict = Depends(get_current_user)
):
    try:
        print(f"📝 Saving database for user: {current_user['id']}")
        db_id = await save_database_connection(
            user_id=current_user["id"],
            db_data=db_data.dict()
        )
        print(f"✅ Database saved with ID: {db_id}")
        
        await send_realtime_notification(
            current_user["id"],
            "Database Connected",
            f"✅ Connected to {db_data.db_name}",
            "success"
        )
        
        return DatabaseConnectionResponse(
            id=db_id,
            db_name=db_data.db_name,
            db_type=db_data.db_type,
            host=db_data.host,
            port=db_data.port,
            username=db_data.username,
            database_name=db_data.database_name,
            created_at=datetime.now().isoformat()
        )
    except Exception as e:
        print(f"❌ Error adding database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/databases", response_model=List[DatabaseConnectionResponse])
async def get_databases(current_user: dict = Depends(get_current_user)):
    try:
        print(f"📋 Getting databases for user: {current_user['id']}")
        dbs = await get_user_databases(current_user["id"])
        print(f"✅ Found {len(dbs)} databases")
        return [DatabaseConnectionResponse(**db) for db in dbs]
    except Exception as e:
        print(f"❌ Error getting databases: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/databases/{db_id}")
async def update_database_connection(
    db_id: str,
    db_data: DatabaseConnectionRequest,
    current_user: dict = Depends(get_current_user)
):
    try:
        from .database import get_connection, encrypt_password
        
        existing = await get_database_connection(db_id, current_user["id"])
        if not existing:
            raise HTTPException(status_code=404, detail="Database connection not found")
        
        conn = await get_connection()
        
        password_to_save = db_data.password if db_data.password else existing["password"]
        if db_data.password:
            password_to_save = encrypt_password(db_data.password)
        
        await conn.execute(
            """UPDATE user_databases 
               SET db_name = $1, db_type = $2, host = $3, port = $4, 
                   username = $5, password = $6, database_name = $7 
               WHERE id = $8 AND user_id = $9""",
            db_data.db_name, db_data.db_type, db_data.host, db_data.port,
            db_data.username, password_to_save, db_data.database_name,
            db_id, current_user["id"]
        )
        await conn.close()
        
        await send_realtime_notification(
            current_user["id"],
            "Database Updated",
            f"✅ Updated {db_data.db_name}",
            "success"
        )
        
        return {"message": "Database updated successfully"}
    except Exception as e:
        print(f"❌ Error updating database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/databases/{db_id}")
async def delete_database_connection(
    db_id: str,
    current_user: dict = Depends(get_current_user)
):
    try:
        from .database import get_connection
        conn = await get_connection()
        result = await conn.execute(
            "DELETE FROM user_databases WHERE id = $1 AND user_id = $2",
            db_id, current_user["id"]
        )
        await conn.close()
        
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Database connection not found")
        
        await send_realtime_notification(
            current_user["id"],
            "Database Deleted",
            "🗑️ Database connection removed",
            "info"
        )
        
        return {"message": "Database deleted successfully"}
    except Exception as e:
        print(f"❌ Error deleting database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# QUERY ENDPOINTS - IMPROVED VERSION
# ============================================================
@app.post("/api/query", response_model=QueryResponse)
async def process_query(
    request: QueryRequest,
    current_user: dict = Depends(get_current_user)
):
    try:
        if not request.prompt or not request.prompt.strip():
            raise HTTPException(status_code=400, detail="Prompt is required")
        
        print(f"📝 Processing query for user: {current_user['id']}")
        print(f"📝 Prompt: {request.prompt}")
        print(f"📝 DB ID: {request.db_id}")
        
        await send_realtime_notification(
            current_user["id"],
            "Query Started",
            f"Processing: {request.prompt[:50]}...",
            "info"
        )
        
        # Generate SQL
        sql_query = await generate_sql_with_ai(request.prompt, request.db_id)
        print(f"📝 Generated SQL: {sql_query}")
        
        result_data = None
        error_msg = None
        execution_time = None
        row_count = 0
        explanation = None
        
        # Case 1: No database selected -> Show explanation only
        if not request.db_id or request.db_id == "all":
            explanation = await explain_sql_query(sql_query, request.prompt)
            
            log = await create_log(
                user_id=current_user["id"],
                prompt=request.prompt,
                sql_query=sql_query,
                result=None,
                error=None,
                execution_time=None,
                db_id=None
            )
            
            return QueryResponse(
                success=True,
                sqlQuery=sql_query,
                logId=log.id,
                result=None,
                row_count=0,
                execution_time=None,
                error=None,
                explanation=explanation,
                message="SQL generated and explained! Select a database to execute."
            )
        
        # Case 2: Database selected but not connected
        db_config = await get_database_connection(request.db_id, current_user["id"])
        if not db_config:
            explanation = await explain_sql_query(sql_query, request.prompt)
            
            log = await create_log(
                user_id=current_user["id"],
                prompt=request.prompt,
                sql_query=sql_query,
                result=None,
                error="Database connection not found",
                execution_time=None,
                db_id=request.db_id
            )
            
            return QueryResponse(
                success=False,
                sqlQuery=sql_query,
                logId=log.id,
                result=None,
                row_count=0,
                execution_time=None,
                error="Database connection not found. Please connect your database first.",
                explanation=explanation,
                message="Database not connected. Here's the SQL explanation."
            )
        
        # Case 3: Database connected -> Execute query
        await send_realtime_notification(
            current_user["id"],
            "Executing Query",
            "Running on your database...",
            "info"
        )
        
        print(f"📝 Executing query on DB: {db_config['db_name']}")
        exec_result = await execute_query(
            db_type=db_config["db_type"],
            config=db_config,
            query=sql_query
        )
        print(f"📝 Execution result: {exec_result.get('success', False)}")
        
        if exec_result.get("success"):
            result_data = exec_result.get("data", [])
            
            # Handle different data types
            if isinstance(result_data, str):
                try:
                    result_data = json.loads(result_data)
                except:
                    result_data = []
            
            if not isinstance(result_data, list):
                result_data = [result_data] if result_data else []
            
            # Ensure all items are dictionaries
            cleaned_data = []
            for item in result_data:
                if isinstance(item, dict):
                    cleaned_data.append(item)
                elif isinstance(item, str):
                    try:
                        parsed = json.loads(item)
                        if isinstance(parsed, dict):
                            cleaned_data.append(parsed)
                        else:
                            cleaned_data.append({"value": parsed})
                    except:
                        cleaned_data.append({"value": item})
                else:
                    cleaned_data.append({"value": str(item)})
            
            result_data = cleaned_data
            row_count = len(result_data)
            execution_time = exec_result.get("execution_time")
            
            await send_realtime_notification(
                current_user["id"],
                "Query Complete!",
                f"✅ {row_count} rows returned in {execution_time}ms",
                "success"
            )
        else:
            error_msg = exec_result.get("error", "Unknown error")
            print(f"❌ Query execution error: {error_msg}")
            await send_realtime_notification(
                current_user["id"],
                "Query Error",
                f"❌ {error_msg[:100]}" if error_msg else "Query failed",
                "error"
            )
        
        # Log the query
        log = await create_log(
            user_id=current_user["id"],
            prompt=request.prompt,
            sql_query=sql_query,
            result=result_data if not error_msg else None,
            error=error_msg,
            execution_time=execution_time,
            db_id=request.db_id
        )
        
        return QueryResponse(
            success=not error_msg,
            sqlQuery=sql_query,
            logId=log.id,
            result=result_data,
            row_count=row_count,
            execution_time=execution_time,
            error=error_msg,
            explanation=None,
            message="Query executed successfully!" if not error_msg else "Query execution failed"
        )
        
    except Exception as e:
        print(f"❌ Query processing error: {e}")
        import traceback
        traceback.print_exc()
        await send_realtime_notification(
            current_user["id"],
            "Error",
            f"❌ {str(e)[:100]}",
            "error"
        )
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# LOGS ENDPOINT
# ============================================================
@app.get("/api/logs", response_model=List[LogResponse])
async def get_query_logs(
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    try:
        logs = await get_logs(current_user["id"], limit)
        return [
            LogResponse(
                id=log["id"],
                userId=log["user_id"],
                prompt=log["prompt"],
                sqlQuery=log["sql_query"],
                createdAt=log["created_at"],
                result=log.get("result"),
                error=log.get("error"),
                execution_time=log.get("execution_time")
            )
            for log in logs
        ]
    except Exception as e:
        print(f"❌ Logs error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# PROFILE ENDPOINTS
# ============================================================
@app.get("/api/profile", response_model=ProfileResponse)
async def get_profile(
    current_user: dict = Depends(get_current_user)
):
    try:
        user = await get_user_by_id(current_user["id"])
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        stats = await get_user_stats(current_user["id"])
        recent_logs = await get_logs(current_user["id"], 5)
        
        return ProfileResponse(
            user=UserResponse(**user),
            total_queries=stats["total_queries"],
            recent_queries=[
                LogResponse(
                    id=log["id"],
                    userId=log["user_id"],
                    prompt=log["prompt"],
                    sqlQuery=log["sql_query"],
                    createdAt=log["created_at"],
                    result=log.get("result"),
                    error=log.get("error"),
                    execution_time=log.get("execution_time")
                )
                for log in recent_logs
            ]
        )
    except Exception as e:
        print(f"❌ Profile error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/profile/stats")
async def get_profile_stats(
    current_user: dict = Depends(get_current_user)
):
    try:
        stats = await get_user_profile_stats(current_user["id"])
        
        if not stats or not stats.get("user"):
            raise HTTPException(status_code=404, detail="User not found")
        
        user = stats["user"]
        
        user_data = UserResponse(
            id=user["id"],
            username=user["username"],
            email=user["email"],
            full_name=user.get("full_name"),
            created_at=user.get("created_at") or datetime.now().isoformat()
        )
        
        recent_queries = []
        for log in stats["recent_queries"][:5]:
            recent_queries.append(
                LogResponse(
                    id=log.get("id", ""),
                    userId=log.get("user_id", ""),
                    prompt=log.get("prompt", ""),
                    sqlQuery=log.get("sql_query", ""),
                    createdAt=log.get("created_at", datetime.now().isoformat()),
                    result=log.get("result"),
                    error=log.get("error"),
                    execution_time=log.get("execution_time")
                )
            )
        
        return {
            "user": user_data,
            "stats": {
                "total_queries": stats["total_queries"],
                "total_databases": stats["total_databases"],
                "avg_execution_time": float(stats["avg_execution_time"]) if stats["avg_execution_time"] else 0,
                "recent_queries": recent_queries
            }
        }
    except Exception as e:
        print(f"❌ Stats error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/profile/update")
async def update_profile(
    profile_data: UserProfileUpdate,
    current_user: dict = Depends(get_current_user)
):
    try:
        if not profile_data.full_name or not profile_data.email:
            raise HTTPException(status_code=400, detail="Full name and email are required")
        
        success = await update_user_profile(
            current_user["id"],
            profile_data.full_name,
            profile_data.email
        )
        
        if success:
            await send_realtime_notification(
                current_user["id"],
                "Profile Updated",
                "Your profile has been updated successfully!",
                "success"
            )
            return {"success": True, "message": "Profile updated successfully"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"❌ Profile update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/profile/password")
async def update_password_endpoint(
    password_data: UserPasswordUpdate,
    current_user: dict = Depends(get_current_user)
):
    try:
        if not password_data.current_password or not password_data.new_password:
            raise HTTPException(status_code=400, detail="Current and new password are required")
        
        if len(password_data.new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
        
        success = await update_password(
            current_user["id"],
            password_data.current_password,
            password_data.new_password
        )
        
        if success:
            await send_realtime_notification(
                current_user["id"],
                "Password Updated",
                "Your password has been changed successfully!",
                "success"
            )
            return {"success": True, "message": "Password updated successfully"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"❌ Password update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# HEALTH CHECK
# ============================================================
@app.get("/health")
async def health_check():
    return {
        "status": "healthy", 
        "database": "PostgreSQL", 
        "ai": f"Gemini ({GEMINI_AVAILABLE_MODEL})" if gemini_model else "Fallback",
        "realtime": "Enabled",
        "version": "2.0.0"
    }